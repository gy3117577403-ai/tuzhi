from __future__ import annotations

import csv
import json
import shutil
import uuid
from pathlib import Path
from typing import Any

from fastapi import HTTPException, UploadFile

from services.procurement_data_normalizer import normalize_offer_row
from services.procurement_models import ProcurementImportResponse, ProcurementResult, ProcurementSourceCreateRequest
from services.procurement_source_config import now_iso
from services.procurement_source_store import PROCUREMENT_DATA_DIR, create_source


IMPORT_ROOT = Path(__file__).resolve().parents[1] / "outputs" / "procurement_imports"


def _read_csv(path: Path) -> list[dict[str, Any]]:
    for encoding in ("utf-8-sig", "gb18030", "utf-8"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return list(csv.DictReader(handle))
        except UnicodeDecodeError:
            continue
    raise HTTPException(status_code=400, detail="CSV 编码无法识别，请使用 UTF-8 或 GB18030")


def _read_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # pragma: no cover - depends on local optional package
        raise HTTPException(status_code=400, detail="当前环境缺少 openpyxl，无法导入 Excel") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    records: list[dict[str, Any]] = []
    for values in rows[1:]:
        records.append({headers[index]: value for index, value in enumerate(values) if index < len(headers)})
    return records


def _read_rows(path: Path) -> tuple[list[dict[str, Any]], str]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path), "csv_upload"
    if suffix in {".xlsx", ".xlsm"}:
        return _read_xlsx(path), "excel_upload"
    raise HTTPException(status_code=400, detail="仅支持 CSV 或 XLSX 报价表")


def load_imported_offers() -> list[ProcurementResult]:
    if not IMPORT_ROOT.exists():
        return []
    offers: list[ProcurementResult] = []
    for file in IMPORT_ROOT.glob("*/offers.json"):
        try:
            raw = json.loads(file.read_text(encoding="utf-8"))
            offers.extend(ProcurementResult(**item) for item in raw)
        except Exception:
            continue
    return offers


async def import_procurement_file(file: UploadFile, source_name: str, platform_label: str) -> ProcurementImportResponse:
    filename = Path(file.filename or "procurement_import.csv").name
    if not filename.lower().endswith((".csv", ".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 CSV 或 XLSX 报价表")

    import_id = uuid.uuid4().hex
    output_dir = IMPORT_ROOT / import_id
    output_dir.mkdir(parents=True, exist_ok=True)
    saved_path = output_dir / filename
    with saved_path.open("wb") as handle:
        shutil.copyfileobj(file.file, handle)

    rows, source_type = _read_rows(saved_path)
    source = create_source(
        ProcurementSourceCreateRequest(
            source_name=source_name or f"报价表 {import_id[:8]}",
            source_type=source_type,  # type: ignore[arg-type]
            enabled=True,
            priority=10,
            platform_label=platform_label or "其他",
            notes="采购手动导入报价表；原始文件保存在本地 outputs，不进入 Git。",
            auth_mode="none",
            safe_mode=True,
        )
    )

    offers: list[ProcurementResult] = []
    warnings: list[str] = []
    for index, row in enumerate(rows, start=2):
        offer, warning = normalize_offer_row(
            row,
            row_index=index,
            source_id=source.source_id,
            source_name=source.source_name,
            source_type=source_type,
            platform_label=str(source.platform_label),
            import_id=import_id,
        )
        if offer:
            offers.append(offer)
        if warning:
            warnings.append(warning)

    (output_dir / "offers.json").write_text(
        json.dumps([offer.model_dump() for offer in offers], ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (output_dir / "import_manifest.json").write_text(
        json.dumps(
            {
                "import_id": import_id,
                "source_id": source.source_id,
                "source_name": source.source_name,
                "source_type": source_type,
                "original_filename": filename,
                "created_at": now_iso(),
                "rows_total": len(rows),
                "rows_imported": len(offers),
                "rows_skipped": len(rows) - len(offers),
            },
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )
    PROCUREMENT_DATA_DIR.mkdir(parents=True, exist_ok=True)

    return ProcurementImportResponse(
        import_id=import_id,
        source_name=source.source_name,
        source_type=source_type,  # type: ignore[arg-type]
        rows_total=len(rows),
        rows_imported=len(offers),
        rows_skipped=len(rows) - len(offers),
        warnings=warnings,
        offers=offers[:20],
    )
